"""
This program was developed on July 25, 2022 by Abhro Chowdhury, with help from Ritvik Das. In order to run the program, first ensure that all installs are in place and that a virtual environment has been created.
Next, any users will need to ensure that they have the directories from line 24 and 25 correct. Once these prerequisites are completed, please ensure that the "EXCEL_TIMESHEETS" folder contains all 
of the timesheets to use (Please note that you will need to copy and paste everyone's timesheets in to this folder before you can run the program). The program should then take a couple minutes to run.
The "EXCEL_OVERTIME" folder contains the banked overtime files for each employee, and here you will find the ouput of the program. The program is designed to write on top of any pre-existing data
on the excel sheets, hence you can run the program multiple times without worrying about previous month's overtime hours disappearing. In terms of upkeep, any new hire's name will need to be added 
to the "employees.py" file. To do this you can simply just type the name into the employee_list variable. 
If any further clarification or help is required, please contact: abhro.chowdhury@caneil.com --or-- abhrajyo@ualberta.ca
"""

# imports 
import os
from pathlib import Path
from re import T
from tkinter import END, Y
from tkinter.font import names
from types import NoneType
import webbrowser
from numpy import full, true_divide
import pandas as pd
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
import glob
from employees import employee_list

"""
Initializations
"""
excel_input_directory = r"C:\Users\caneil7\Desktop\timesheet testing\EXCEL_TIMESHEETS"
excel_output_directory = r"C:\Users\caneil7\Desktop\timesheet testing\EXCEL_OVERTIME"

timesheet_month = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]  # Array of months for input file specification.
outputsheet_month = ["JAN 22", "FEB 22", "MAR 22", "APR 22", "MAY 22", "JUN 22", "JUL 22", "AUG 22", "SEP 22", "OCT 22", "NOV 22", "DEC 22"]    # Array of months for ouput file specification.
full_month = ["January", "February", "March", "April", "May", "June", "July", "August", "September", "October", "November", "December"]     # Array of full months to paste onto output template.
print("The program is initializing...")

"""
ALL FUNCTIONS
"""

# This is the main part of the code; This function will pull how many overtime hours is worked per day and then store that data.
# The function itself is complete and working, however based on how we can pull up the other excel files, we'll need to find
# a way to then store that data for the third step

def OTperday(specialcolumn):
    mylist = []    # creating an empty list to later append values to 
    for i in range(11, 30):    # range of rows that I want to iterate over (usually 11 to 70)
        checktype = type(month.cell(row=i, column=specialcolumn).value)  # this will pick out the specific cell, and check the type of value in it
        if checktype == int:   # I want to only add integers to my list
            data = month.cell(row=i, column=specialcolumn).value     # gives me value for the column we want, and will iterate over each row in that column
            mylist.append(data)    # adding values into the list, this will give us the final overtime hours in that particular day
    thesum = sum(mylist)    # taking total of the list to calculate total OT for that particular day
    return thesum


# This function will populate the dates for the sheet we are working on 

def write_row(write_sheet, row_num: int, starting_column: int, write_values: list):    # row and column will always stay same, just change the list to match month
    for i, value in enumerate(write_values):    # loop over each row in the same column 
        write_sheet.cell(row_num + i, starting_column, value)     # writes to each row in the column


#Function copying and writing all days in a month down onto output template

def calendarCopy(i, y):
    output_ws = output_workbook[outputsheet_month[y]]
    Cal_list = []
    for day in range(1, i):
        dayString = str(day)
        calendar = (str(full_month[y]) + " " + dayString + ", 2022")
        Cal_list.append(calendar)
        write_row(output_ws, 12, 1, Cal_list)

"""
MAIN SCRIPT
"""
print("The program is running...")

z = 0
while z < len(employee_list):
    name = employee_list[z]
    for filename in os.listdir(excel_input_directory) and os.listdir(excel_output_directory):
        if filename.endswith(".xlsx"):
            timesheet_wb = load_workbook(f"./EXCEL_TIMESHEETS/TS-2022_{name}.xlsx")   # loading excel book
            ws = timesheet_wb.active   # loading excel sheet
            output_file = f"./EXCEL_OVERTIME/BankedOT_{name}.xlsx"    # saving the output file name for later
            output_workbook = load_workbook(output_file)   # loading in the output file
            break

    # To loop over initial input excel document in order to copy necessary data and paste onto template sheet.
    x = 0
    while x < len(timesheet_month):
        output_ws = output_workbook[outputsheet_month[x]]
        month = timesheet_wb[timesheet_month[x]]
        OTlist = []
        for d in range(1, 31):
            OTdata = OTperday(4 + d*2)
            OTlist.append(OTdata)
            write_row(output_ws, 12, 2, OTlist)
        x += 1

    # To fill in excel sheet with 
    y = 0
    while y < len(full_month):
        if y == 1:
            calendarCopy(29, y)
        elif y % 2 == 1 and y < 7:
            calendarCopy(31, y)
        elif y % 2 == 0 and y >= 7:
            calendarCopy(31, y)
        else:
            calendarCopy(32, y)
        y += 1
    
    output_workbook.save(output_file)
    z += 1

print("The program has finished. The updated banked overtime sheets can be found in 'EXCEL_OVERTIME'.")
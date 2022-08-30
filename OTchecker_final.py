import openpyxl
from openpyxl import load_workbook
import os
from pathlib import Path
from re import T
from tkinter import END, Y
from tkinter.font import names
from types import NoneType
import webbrowser
from numpy import full, true_divide
import pandas as pd
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
import glob
from employees_short import employee_list


outputsheet_month = ["JAN 22", "FEB 22", "MAR 22", "APR 22", "MAY 22", "JUN 22", "JUL 22", "AUG 22", "SEP 22", "OCT 22", "NOV 22", "DEC 22"]    # Array of months for ouput file specification.

# excel_input_directory = r"C:\Users\caneil7\Desktop\OTIS_shortform\TIMESHEETS_SHORT"
excel_output_directory = r"C:\Users\caneil7\Desktop\OTIS_shortform\OVERTIME_SHORT"


# function to write rows
def write_row(write_sheet, row_num: int, starting_column: int, write_values: list):    # row and column will always stay same, just change the list to match month
    for i, value in enumerate(write_values):    # loop over each row in the same column 
        write_sheet.cell(row_num + i, starting_column, value)     # writes to each row in the column

# looping over everyone in the list

outputsheet_month = input("Please enter the month you are checking for (FORMAT is like {DEC 22}: ")    # Haley must enter this too, format is: {MONTH 22}
print("Processing...")
z = 0

while z < len(employee_list):
    hourslist = []
    name = employee_list[z]
    for filename in  os.listdir(excel_output_directory):
        if filename.endswith(".xlsx"):
            output_file = f"./OVERTIME_SHORT/BankedOT_{name}.xlsx"    # saving the output file name for later 
            output_workbook = load_workbook(output_file, data_only=True)   # loading in the output file
            output_ws = output_workbook[outputsheet_month]  # loading in the month haley chooses
            break
    for row in range(12, 42):  
        for column in "B":  #Here you can add or reduce the columns
            cell_name = "{}{}".format(column, row)
            values1 = output_ws[cell_name].value # the value of the specific cell
            hourslist.append(values1)

    total = sum(hourslist)
    if total != 0:
        print(f"{name} has {total} overtime hours for {outputsheet_month}")   

    z += 1

